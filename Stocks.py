#-------------------------------------------------------------------------------
# Name:        Stocks
# Purpose:     Create an XML workbook to record a bunch of stocks
#
# Author:      John Eichenberger
#
# Created:     11/06/2021
# Copyright:   (c) John Eichenberger 2021
# Licence:     GNU GENERAL PUBLIC LICENSE Version 3, 29 June 2007
#-------------------------------------------------------------------------------

# !pip install yahoofinancials
# !pip install openpyxl

# not using, but also considered
# !pip install yfinance
# !pip install matplotlib

from yahoofinancials import YahooFinancials as yf
import glob, re, os, sys, time
from math import log, floor
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors
from datetime import datetime, timedelta

# Keep track of all fields for which data is collected
summary_fields = []

def fmt_cell(number, percent=''):
    """ Get fancy how cells are formatted, specifically numbers.
        Format integers less than 1,000 quite simply.
        Format fractions with exactly two decimal points.
        Display negative numbers in red.
        Use scientific notation on all numbers bigger than 1,000,
        using a power of 3.  That way one can translate to thousands, millions,
        billions, etc. in their heads.
    """
    if isinstance(number, int):
        fmt = '0'+percent       # Default format for integers
    elif isinstance(number, float):
        fmt = '0.00'+percent    # Default format for real numbers
    else:
        return 'General'

    if number < 0:
        number = -number    # log function requires positive numbers
    elif number == 0:
        number = 1          # log function requires positive numbers
    pwr = log(number, 1000)
    if pwr >= 1:
        fmt = '##0.00E+0'[2-int((pwr % 1) * 3):]
    return fmt + ';[RED]-' + fmt

def GetData(ticker):
    """ Retrieve stock data and add it to a dictionary entry """
    global summary_fields

    # Return None for any ticker that has no data
    stock = yf(ticker)
    data = stock.get_stock_price_data()[ticker]
    if data == None:
        print("{} not found".format(ticker))
        return data

    # Pull select data from get_stock_price_data
    fields = {}
    fields['Symbol']= 'symbol'                          # A1
    fields['Exchange'] = 'exchangeName'                 # B1
    fields['Type'] = 'quoteType'                        # C1
    fields['Name'] = 'shortName'                        # D1
    fields['Price'] = 'regularMarketPrice'              # E1
    rd = {}
    for fld in fields:
        if not fld in summary_fields:
            summary_fields += [fld]
        rd[fld] = data[fields[fld]]
        print("{} = {}".format(fld, data[fields[fld]]))

    # Pull select data from get_summary_data()
	# Each field could be zero or None.
    data = stock.get_summary_data()[ticker]
    if data is not None:
        fields = {}
        fields['50 Day'] = 'fiftyDayAverage'                # F1
        fields['200 Day'] = 'twoHundredDayAverage'		    # G1
        fields['52wk High'] = 'fiftyTwoWeekHigh'			# H1
        fields['52wk Low'] = 'fiftyTwoWeekLow'				# I1
        fields['CAP'] = 'marketCap'                         # J1
        fields['Volume'] = 'averageVolume'		           	# K1
        fields['10 Day'] = 'averageVolume10days'		    # L1
        fields['PE ->'] = 'forwardPE'                       # M1
        fields['PE <-'] = 'trailingPE'				        # N1
        for fld in fields:
            if not fld in summary_fields:
                summary_fields += [fld]
            if fields[fld] in data:
                val = data[fields[fld]]
                if val == 0:
                    val = None  # Make zeros disappear
                rd[fld] = val
                print("{} = {}".format(fld, val))

    # Pull select data from get_key_statistics_data()
	# Could be zero or None.
    data = stock.get_key_statistics_data()[ticker]
    if data is not None:
        fields = {}
        fields['Eps ->'] = 'forwardEps'				        # O1
        fields['Eps <-'] = 'trailingEps'				    # P1
        fields['Beta'] = 'beta'                             # Q1
        fields['Book'] = 'bookValue'						# R1
        fields['Price/Book'] = 'priceToBook'				# S1
        fields['Outstanding'] = 'sharesOutstanding'			# T1
        fields['Float'] = 'floatShares'						# U1
        fields['Short%'] = 'sharesPercentSharesOut'			# V1
        fields['Insiders%'] = 'heldPercentInsiders'			# W1
        fields['Institutions%'] = 'heldPercentInstitutions'	# X1
        fields['Dividend'] = 'lastDividendValue'			# Y1
        for fld in fields:
            if not fld in summary_fields:
                summary_fields += [fld]
            if fields[fld] in data:
                val = data[fields[fld]]
                if val == 0:
                    val = None  # Make zeros disappear
                rd[fld] = val
                print("{} = {}".format(fld, val))

    # Pull select data from get_stock_earnings_data()
    data = stock.get_stock_earnings_data()
    if data is not None:
        if data[ticker] == None:
            print("No earnings data for {}".format(ticker))
        else:
            key = 'EPS'
            if not 'earningsData' in data[ticker]:
                print("No earnings for {}".format(ticker))
            else:
                print("Recording {} data".format(key))
                rd[key] = {}
                earnings = data[ticker]['earningsData']
                for q in earnings['quarterly']:
                    eps = {}
                    eps['actual'] = q['actual']
                    eps['estimate'] = q['estimate']
                    rd[key][q['date']] = eps
                if 'currentQuarterEstimate' in earnings:
                    try: # Create the name of the current quarter EPS value
                        currentQ = earnings['currentQuarterEstimateDate'] + \
                            str(earnings['currentQuarterEstimateYear'])
                    except:
                        currentQ = "NextQ"
                    rd[key][currentQ] = {}
                    rd[key][currentQ]['estimate'] = earnings['currentQuarterEstimate']

            key = 'Revenue'
            if not 'financialsData' in data[ticker]:
                print("No financial data for {}".format(ticker))
            else:
                print("Recording {} data".format(key))
                rd[key] = {}
                financials = data[ticker]['financialsData']
                for q in financials['yearly']:
                    rd[key][q['date']] = q['revenue']
                for q in financials['quarterly']:
                    rd[key][q['date']] = q['revenue']

                key = 'Earnings'
                print("Recording {} data".format(key))
                rd[key] = {}
                for q in financials['yearly']:
                    rd[key][q['date']] = q['earnings']
                for q in financials['quarterly']:
                    rd[key][q['date']] = q['earnings']

    # Pull select data from get_financial_stmts('quarterly', 'income')
    data = stock.get_financial_stmts('quarterly', 'income')['incomeStatementHistoryQuarterly'][ticker]
    if data is not None:
        key = 'Income'
        if data == None:
            print("No financial {} statement for {}".format(key, ticker))
        else:
            print("Recording {} data".format(key))
            fields = ['totalOperatingExpenses', 'totalOtherIncomeExpenseNet', 'totalRevenue']
            rd[key] = {}
            for list in data:
                for day in list:
                    rd[key][day] = {}
                    for fld in fields:
                        if fld in list[day]:
                            rd[key][day][fld] = list[day][fld]

    # Pull select data from get_financial_stmts('quarterly', 'balance')
    data = stock.get_financial_stmts('quarterly', 'balance')['balanceSheetHistoryQuarterly'][ticker]
    if data is not None:
        key = 'Balance Sheet'
        if data == None:
            print("No financial {} statement for {}".format(key, ticker))
        else:
            print("Recording {} data".format(key))
            fields = ['totalAssets', 'totalCurrentAssets', 'totalCurrentLiabilities',
                    'totalLiab', 'totalStockholderEquity']
            rd[key] = {}
            for list in data:
                for day in list:
                    rd[key][day] = {}
                    for fld in fields:
                        if fld in list[day]:
                            rd[key][day][fld] = list[day][fld]

    # Pull select data from get_financial_stmts('quarterly', 'cash')
    data = stock.get_financial_stmts('quarterly', 'cash')['cashflowStatementHistoryQuarterly'][ticker]
    if data is not None:
        key = 'Cash Flow'
        if data == None:
            print("No financial {} statement for {}".format(key, ticker))
        else:
            print("Recording {} data".format(key))
            fields = ['totalCashFromFinancingActivities',
                    'totalCashFromOperatingActivities',
                    'totalCashflowsFromInvestingActivities']
            rd[key] = {}
            for list in data:
                for day in list:
                    rd[key][day] = {}
                    for fld in fields:
                        if fld in list[day]:
                            rd[key][day][fld] = list[day][fld]

    # Pull select data from get_historical_price_data(start_date, end_date, 'daily')
    start_date = (datetime.today() - timedelta(days=122)).strftime("%Y-%m-%d")
    end_date = datetime.today().strftime("%Y-%m-%d")
    try:
        data = stock.get_historical_price_data(start_date, end_date, 'daily')[ticker]['prices']
        key = 'History'
        if data == None:
            print("No {} statement for {}".format(key, ticker))
        else:
            print("Recording {} data".format(key))
            fields = ['open', 'close', 'volume', 'high', 'low', 'volume', 'adjclose']
            rd[key] = {}
            for entry in data:
                day = entry['formatted_date']
                rd[key][day] = {}
                for fld in fields:
                    rd[key][day][fld] = entry[fld]
    except:
        pass

    return rd


#-------------------------------------------------------------------------------
if __name__ == '__main__':
    # Process command line arguments, each is a stock ticker
    # Create a dictionary with all of the gathered data
    stocks = {}
    for arg in sys.argv[1:]:
        # Skip any ticker that has no data
        print("{}:".format(arg.upper()))
        data = GetData(arg.upper())
        if data == None:
            print("not found")
        else:
            # Create a new ticker and a worksheet for that ticker
            stocks[arg.upper()] = data
            print()

    # If nothing was found, don't create a workbook
    if len(stocks) == 0:
        exit()

    # Use the current folder unless "Playground" is defined as an environment variable.
    folder = os.environ.get('Playground')
    if folder is not None:
        os.chdir(folder)

    # Create or replace an xml workbook
    wb = Workbook()
    wb_filename = "Stocks.xml"
    summary = wb.active
    summary.title = "Summary"
    summary = wb.active

    # Create column headers for the summary workbook
    col = 1
    for field in summary_fields:
        summary.cell(1, col, field)
        summary.cell(1, col).font = Font(bold=True)
        col += 1

    # Process the gathered data
    for ticker in stocks:

        # Add summary data items
        row = len(wb.worksheets)*2
        col = 1
        for field in summary_fields:
            data = stocks[ticker]
            if field in data:
                if not isinstance(data[field], dict):
                    # Add percentage format to any field with '%' in the heading
                    if '%' in field:
                        percent = '%'
                    else:
                        percent = ''

                    # Add the summary data itself
                    summary.cell(row, col, data[field]).number_format = fmt_cell(data[field], percent)

            col += 1

        # Add a line of math
        price = '$E'+str(row)
        row += 1
        alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        for col in range(6, 10): # e.g. '=$E2/F2-1' as a percentage
            summary.cell(row, col, '='+price+'/'+alphabet[col-1]+str(row-1)+'-1')
            summary.cell(row, col).number_format = '+0.0%;[RED]-0.0%'
        for col in range(12, 13): # e.g. '=L2/K2-1' as a percentage
            summary.cell(row, col, '='+alphabet[col-1]+str(row-1)+'/'+alphabet[col-2]+str(row-1)+'-1')
            summary.cell(row, col).number_format = '+0.0%;[RED]-0.0%'
        for col in range(21, 22): # e.g. '=U2/T2' as a percentage
            summary.cell(row, col, '='+alphabet[col-1]+str(row-1)+'/'+alphabet[col-2]+str(row-1))
            summary.cell(row, col).number_format = '+0.0%;[RED]-0.0%'

        # Create a new worksheet for each ticker
        sheet = wb.create_sheet(ticker)

        # Widen each column of stock specific sheets
        for letter in alphabet:
            sheet.column_dimensions[letter].width = 10.0

        # Add worksheet data items, starting with EPS
        if 'EPS' in data:
            sheet['A1'] = "EPS"
            sheet.merge_cells('A1:C1')
            sheet['A1'].alignment = Alignment(horizontal='center')
            sheet['A1'].font = Font(bold=True)
            sheet['B2'] = "Actual"
            sheet['B2'].font = Font(bold=True)
            sheet['C2'] = "Estimate"
            sheet['C2'].font = Font(bold=True)
            row = 3
            for period in data['EPS']:
                sheet.cell(row,1,period)
                if 'actual' in data['EPS'][period]:
                    sheet.cell(row, 2, data['EPS'][period]['actual'])
                    sheet.cell(row, 2).number_format = fmt_cell(9.99)
                    sheet.cell(row, 3, "=" + sheet.cell(row,2).coordinate + "-" + str(data['EPS'][period]['estimate']))
                    sheet.cell(row, 3).number_format = fmt_cell(9.99)
                else:
                    sheet.cell(row, 3, data['EPS'][period]['estimate'])
                    sheet.cell(row, 3).number_format = fmt_cell(9.99)
                row += 1

        # Add Earnings and Revenue
        col = 5
        for choice in ["Revenue", 'Earnings']:
            if choice in data:
                row = 1
                sheet.cell(row, col, choice)
                sheet.cell(row, col).font = Font(bold=True)
                row += 1
                for period in data[choice]:
                    sheet.cell(row, col, period)
                    sheet.cell(row, col+1, data[choice][period]).number_format = fmt_cell(data[choice][period])
                    row += 1
                col += 3
        if sheet.max_row > 1:
            sheet.merge_cells('E1:F1')
            sheet.merge_cells('H1:I1')
            sheet['E1'].alignment = Alignment(horizontal='center')
            sheet['H1'].alignment = Alignment(horizontal='center')

        # Add Income
        key = 'Income'
        if key in data:
            columns = {
                'Revenue': 'totalRevenue',
                'Expenses': 'totalOperatingExpenses',
                'Other Income': 'totalOtherIncomeExpenseNet'
                }
            row = sheet.max_row + 3
            if row == 4:
                row = 1
            col = 1
            sheet.cell(row, col, key)   # Create header row
            sheet.cell(row, col).font = Font(bold=True)
            for fld in columns:
                col += 2
                sheet.cell(row, col, fld)
                sheet.cell(row, col).font = Font(bold=True)
            for date in data[key]:
                row += 1
                col = 1
                for fld in columns:
                    col += 2
                    if columns[fld] in data[key][date]:
                        val = data[key][date][columns[fld]]
                        sheet.cell(row, col, val).number_format = fmt_cell(val)
                dttm = datetime.strptime(date, "%Y-%m-%d")
                sheet.cell(row, col+2, dttm).number_format = 'MM/DD/YY'

        # Add Balance Sheet
        key = 'Balance Sheet'
        if key in data:
            columns = {
                'Total Assets': 'totalAssets',
                'Total Liabilities': 'totalLiab',
                'Total Equity': 'totalStockholderEquity',
                'Current Assets': 'totalCurrentAssets',
                'Current Liabilities': 'totalCurrentLiabilities'
                }
            row = sheet.max_row + 3
            if row == 4:
                row = 1
            col = 1
            sheet.cell(row, col, key)   # Create header row
            sheet.cell(row, col).font = Font(bold=True)
            for fld in columns:
                col += 2
                sheet.cell(row, col, fld)
                sheet.cell(row, col).font = Font(bold=True)
            for date in data[key]:
                row += 1
                col = 1
                for fld in columns:
                    col += 2
                    if columns[fld] in data[key][date]:
                        val = data[key][date][columns[fld]]
                        sheet.cell(row, col, val).number_format = fmt_cell(val)
                dttm = datetime.strptime(date, "%Y-%m-%d")
                sheet.cell(row, col+2, dttm).number_format = 'MM/DD/YY'

        # Add Cash Flow
        key = 'Cash Flow'
        if key in data:
            columns = {
                'Financing': 'totalCashFromFinancingActivities',
                'Operating': 'totalCashFromOperatingActivities',
                'Investing': 'totalCashflowsFromInvestingActivities'
                }
            row = sheet.max_row + 3
            if row == 4:
                row = 1
            col = 1
            sheet.cell(row, col, key)   # Create header row
            sheet.cell(row, col).font = Font(bold=True)
            for fld in columns:
                col += 2
                sheet.cell(row, col, fld)
                sheet.cell(row, col).font = Font(bold=True)
            for date in data[key]:
                row += 1
                col = 1
                for fld in columns:
                    col += 2
                    if columns[fld] in data[key][date]:
                        val = data[key][date][columns[fld]]
                        sheet.cell(row, col, val).number_format = fmt_cell(val)
                dttm = datetime.strptime(date, "%Y-%m-%d")
                sheet.cell(row, col+2, dttm).number_format = 'MM/DD/YY'

        # Add History
        key = 'History'
        if key in data:
            columns = {
                'Open': 'open',
                'Close': 'close',
                'Volume': 'volume',
                'High': 'high',
                'Low': 'low',
                'Adjusted close': 'adjclose'
                }
            row = sheet.max_row + 3
            if row == 4:
                row = 1
            col = 1
            sheet.cell(row, col, key)   # Create header row
            sheet.cell(row, col).font = Font(bold=True)
            for fld in columns:
                col += 1
                sheet.cell(row, col, fld)
                sheet.cell(row, col).font = Font(bold=True)
            for date in data[key]:
                row += 1
                col = 1
                for fld in columns:
                    col += 1
                    if columns[fld] in data[key][date]:
                        val = data[key][date][columns[fld]]
                        sheet.cell(row, col, val).number_format = fmt_cell(val)
                dttm = datetime.strptime(date, "%Y-%m-%d")
                sheet.cell(row, col+2, dttm).number_format = 'MM/DD/YY'

    # Save the results
    summary.freeze_panes = 'B2'
    summary.auto_filter.ref = "A1:Z"+str(summary.max_row+1)

    # Automate column widths
    # Default all columns to the same width
    for col in summary.columns:
        letter = col[0].column_letter
        summary.column_dimensions[letter].width = 10.0
    summary.column_dimensions['D'].width = 32.0    # Name

    try:
        wb.save(wb_filename)
        wb.close()
    except:
        print("Is the output file ({}) open?".format(wb_filename))

